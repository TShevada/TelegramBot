import logging
import asyncio
import openpyxl
import os
from datetime import datetime
from aiogram import Bot, Dispatcher, types, F
from aiogram.enums import ParseMode
from aiogram.client.default import DefaultBotProperties
from aiogram.filters import Command
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup
from collections import defaultdict

# --- Configuration ---
TOKEN = "7883966462:AAG2udLydnyXDibLWyw8WrlVntzUB-KMXfE"
YOUR_TELEGRAM_ID = 1291104906
PHOTOS_DIR = "payment_screenshots"
WELCOME_BANNER = "welcome_banner.jpg"
PAYMENT_CARD = "4169 7388 9268 3164"

# --- Setup ---
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
os.makedirs(PHOTOS_DIR, exist_ok=True)

bot = Bot(token=TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.MARKDOWN))
dp = Dispatcher()

# --- Storage ---
user_lang = {}
user_data = {}
save_counter = defaultdict(int)
admin_pending_actions = {}
pending_approvals = {}  # Для хранения заявок на модерацию

# --- Ticket Prices ---
TICKET_TYPES = {
    "standard": {
        "ru": {"name": "Стандарт", "price": "20 манат", "desc": "включает welcome cocktails (безалкогольные)"},
        "az": {"name": "Standart", "price": "20 manat", "desc": "welcome cocktails (alkogolsuz) daxildir"},
        "en": {"name": "Standard", "price": "20 AZN", "desc": "includes welcome cocktails (non-alcohol)"}
    },
    "vip": {
        "ru": {"name": "VIP", "price": "40 манат", "desc": "можно потратить 20 манат на еду и напитки"},
        "az": {"name": "VIP", "price": "40 manat", "desc": "20 manatı yemək və içkilərə xərcləmək olar"},
        "en": {"name": "VIP", "price": "40 AZN", "desc": "20 AZN can be spent on food and drinks"}
    },
    "exclusive": {
        "ru": {"name": "Эксклюзив", "price": "60 манат", "desc": "можно потратить 30 манат на еду и напитки"},
        "az": {"name": "Eksklüziv", "price": "60 manat", "desc": "30 manatı yemək və içkilərə xərcləmək olar"},
        "en": {"name": "Exclusive", "price": "60 AZN", "desc": "30 AZN can be spent on food and drinks"}
    }
}

# --- Helper Functions ---
def is_admin(user_id: int) -> bool:
    return user_id == YOUR_TELEGRAM_ID

def get_lang_keyboard():
    return types.ReplyKeyboardMarkup(
        keyboard=[
            [types.KeyboardButton(text="🇷🇺 Русский"), types.KeyboardButton(text="🇦🇿 Azərbaycan")],
            [types.KeyboardButton(text="🇬🇧 English")]
        ],
        resize_keyboard=True
    )

def get_menu_keyboard(lang):
    if lang == "ru":
        return types.ReplyKeyboardMarkup(
            keyboard=[
                [types.KeyboardButton(text="🎫 Билеты")],
                [types.KeyboardButton(text="📅 Ближайшие события")],
                [types.KeyboardButton(text="📞 Контакты")],
                [types.KeyboardButton(text="🌐 Сменить язык")]
            ],
            resize_keyboard=True
        )
    elif lang == "az":
        return types.ReplyKeyboardMarkup(
            keyboard=[
                [types.KeyboardButton(text="🎫 Biletlər")],
                [types.KeyboardButton(text="📅 Yaxın tədbirlər")],
                [types.KeyboardButton(text="📞 Əlaqə")],
                [types.KeyboardButton(text="🌐 Dil dəyiş")]
            ],
            resize_keyboard=True
        )
    else:  # English
        return types.ReplyKeyboardMarkup(
            keyboard=[
                [types.KeyboardButton(text="🎫 Tickets")],
                [types.KeyboardButton(text="📅 Upcoming events")],
                [types.KeyboardButton(text="📞 Contacts")],
                [types.KeyboardButton(text="🌐 Change language")]
            ],
            resize_keyboard=True
        )

def get_ticket_type_keyboard(lang):
    if lang == "ru":
        return types.ReplyKeyboardMarkup(
            keyboard=[
                [types.KeyboardButton(text="Стандарт (20 манат)")],
                [types.KeyboardButton(text="VIP (40 манат)")],
                [types.KeyboardButton(text="Эксклюзив (60 манат)")],
                [types.KeyboardButton(text="⬅️ Назад")]
            ],
            resize_keyboard=True
        )
    elif lang == "az":
        return types.ReplyKeyboardMarkup(
            keyboard=[
                [types.KeyboardButton(text="Standart (20 manat)")],
                [types.KeyboardButton(text="VIP (40 manat)")],
                [types.KeyboardButton(text="Eksklüziv (60 manat)")],
                [types.KeyboardButton(text="⬅️ Geri")]
            ],
            resize_keyboard=True
        )
    else:  # English
        return types.ReplyKeyboardMarkup(
            keyboard=[
                [types.KeyboardButton(text="Standard (20 AZN)")],
                [types.KeyboardButton(text="VIP (40 AZN)")],
                [types.KeyboardButton(text="Exclusive (60 AZN)")],
                [types.KeyboardButton(text="⬅️ Back")]
            ],
            resize_keyboard=True
        )

def get_admin_keyboard():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📊 Статистика", callback_data="admin_stats"),
         InlineKeyboardButton(text="📝 Последние заявки", callback_data="admin_last_orders")],
        [InlineKeyboardButton(text="🔍 Поиск по ID", callback_data="admin_search"),
         InlineKeyboardButton(text="🔄 Обновить", callback_data="admin_refresh")]
    ])

async def generate_stats_report():
    try:
        wb = openpyxl.load_workbook("tickets.xlsx")
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        
        total = len(rows) - 1
        if total <= 0:
            return "📭 Нет данных о заявках."
            
        types_count = defaultdict(int)
        for row in rows[1:]:
            types_count[row[3]] += 1
            
        report = (
            f"📈 *Статистика заявок*\n\n"
            f"• Всего заявок: {total}\n"
            f"• Стандарт: {types_count.get('standard', 0)}\n"
            f"• VIP: {types_count.get('vip', 0)}\n"
            f"• Эксклюзив: {types_count.get('exclusive', 0)}\n\n"
            f"Ожидают подтверждения: {len([x for x in pending_approvals.values() if x['approved'] is None])}\n"
            f"Последняя запись:\n"
            f"🕒 {rows[-1][6]}"
        )
        return report
    except Exception as e:
        logger.error(f"Stats error: {e}")
        return f"⚠️ Ошибка генерации отчёта: {e}"

async def get_last_orders(count=5):
    try:
        wb = openpyxl.load_workbook("tickets.xlsx")
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))[-count:]
        
        if len(rows) == 0:
            return "📭 Нет заявок."
            
        report = "📋 *Последние заявки:*\n\n"
        for row in rows:
            report += (
                f"🔹 *ID:* {row[0]}\n"
                f"👤 *{row[1]}*\n"
                f"📞 `{row[2]}`\n"
                f"🎟 {row[3]} ({row[4]})\n"
                f"🕒 {row[6]}\n"
                "━━━━━━━━━━━━━━\n"
            )
        return report
    except Exception as e:
        logger.error(f"Orders error: {e}")
        return f"⚠️ Ошибка: {e}"

# --- Handlers ---
@dp.message(Command("start"))
async def start_cmd(message: types.Message):
    try:
        if os.path.exists(WELCOME_BANNER):
            await message.answer_photo(types.InputFile(WELCOME_BANNER))
    except Exception as e:
        logger.error(f"Banner error: {e}")
    await message.answer("Выберите язык / Select language / Dil seçin:", reply_markup=get_lang_keyboard())

@dp.message(F.text.in_(["🇷🇺 Русский", "🇦🇿 Azərbaycan", "🇬🇧 English"]))
async def set_language(message: types.Message):
    lang_map = {
        "🇷🇺 Русский": "ru",
        "🇦🇿 Azərbaycan": "az",
        "🇬🇧 English": "en"
    }
    user_lang[message.from_user.id] = lang_map[message.text]
    
    confirmation = {
        "ru": "Язык установлен: Русский",
        "az": "Dil seçildi: Azərbaycan",
        "en": "Language set: English"
    }[lang_map[message.text]]
    
    await message.answer(confirmation, reply_markup=get_menu_keyboard(lang_map[message.text]))

@dp.message(F.text.in_(["📅 Ближайшие события", "📅 Yaxın tədbirlər", "📅 Upcoming events"]))
async def events_handler(message: types.Message):
    lang = user_lang.get(message.from_user.id, "en")
    events_info = {
        "ru": "Текущий ивент: Afro-Party в Voodoo!\n"
              "📅 Дата: 27 апреля 2025\n"
              "🕒 Время: 18:00 - 00:00\n"
              "📍 Место: Рестобар Voodoo, ТРЦ Наргиз Молл, 3 этаж",
        "az": "Cari tədbir: Afro-Party Voodoo-da!\n"
              "📅 Tarix: 27 Aprel 2025\n"
              "🕒 Vaxt: 18:00 - 00:00\n"
              "📍 Yer: Voodoo Restobar, Nargiz Mall, 3-cü mərtəbə",
        "en": "Current event: Afro-Party at Voodoo!\n"
              "📅 Date: April 27, 2025\n"
              "🕒 Time: 6:00 PM - 12:00 AM\n"
              "📍 Location: Voodoo Restobar, Nargiz Mall, 3rd floor"
    }[lang]
    await message.answer(events_info)

@dp.message(F.text.in_(["📞 Контакты", "📞 Əlaqə", "📞 Contacts"]))
async def contacts_handler(message: types.Message):
    lang = user_lang.get(message.from_user.id, "en")
    contact_info = {
        "ru": "📞 Контакты:\nТелефон: +994 10 531 24 06",
        "az": "📞 Əlaqə:\nTelefon: +994 10 531 24 06",
        "en": "📞 Contacts:\nPhone: +994 10 531 24 06"
    }[lang]
    await message.answer(contact_info)

@dp.message(F.text.in_(["🌐 Сменить язык", "🌐 Dil dəyiş", "🌐 Change language"]))
async def change_lang_handler(message: types.Message):
    await message.answer(
        "Выберите язык / Select language / Dil seçin:",
        reply_markup=get_lang_keyboard()
    )

@dp.message(F.text.in_(["🎫 Билеты", "🎫 Biletlər", "🎫 Tickets"]))
async def tickets_menu_handler(message: types.Message):
    lang = user_lang.get(message.from_user.id, "en")
    
    tickets_info = {
        "ru": "🎟 Доступные билеты:\n\n"
              f"1. {TICKET_TYPES['standard']['ru']['name']} - {TICKET_TYPES['standard']['ru']['price']}\n"
              f"   {TICKET_TYPES['standard']['ru']['desc']}\n\n"
              f"2. {TICKET_TYPES['vip']['ru']['name']} - {TICKET_TYPES['vip']['ru']['price']}\n"
              f"   {TICKET_TYPES['vip']['ru']['desc']}\n\n"
              f"3. {TICKET_TYPES['exclusive']['ru']['name']} - {TICKET_TYPES['exclusive']['ru']['price']}\n"
              f"   {TICKET_TYPES['exclusive']['ru']['desc']}\n\n"
              "Выберите тип билета:",
        "az": "🎟 Mövcud biletlər:\n\n"
              f"1. {TICKET_TYPES['standard']['az']['name']} - {TICKET_TYPES['standard']['az']['price']}\n"
              f"   {TICKET_TYPES['standard']['az']['desc']}\n\n"
              f"2. {TICKET_TYPES['vip']['az']['name']} - {TICKET_TYPES['vip']['az']['price']}\n"
              f"   {TICKET_TYPES['vip']['az']['desc']}\n\n"
              f"3. {TICKET_TYPES['exclusive']['az']['name']} - {TICKET_TYPES['exclusive']['az']['price']}\n"
              f"   {TICKET_TYPES['exclusive']['az']['desc']}\n\n"
              "Bilet növünü seçin:",
        "en": "🎟 Available tickets:\n\n"
              f"1. {TICKET_TYPES['standard']['en']['name']} - {TICKET_TYPES['standard']['en']['price']}\n"
              f"   {TICKET_TYPES['standard']['en']['desc']}\n\n"
              f"2. {TICKET_TYPES['vip']['en']['name']} - {TICKET_TYPES['vip']['en']['price']}\n"
              f"   {TICKET_TYPES['vip']['en']['desc']}\n\n"
              f"3. {TICKET_TYPES['exclusive']['en']['name']} - {TICKET_TYPES['exclusive']['en']['price']}\n"
              f"   {TICKET_TYPES['exclusive']['en']['desc']}\n\n"
              "Select ticket type:"
    }[lang]
    
    await message.answer(tickets_info, reply_markup=get_ticket_type_keyboard(lang))

@dp.message(F.text.in_(["⬅️ Назад", "⬅️ Geri", "⬅️ Back"]))
async def back_handler(message: types.Message):
    lang = user_lang.get(message.from_user.id, "en")
    await message.answer("Главное меню" if lang == "ru" else "Ana menyu" if lang == "az" else "Main menu", 
                        reply_markup=get_menu_keyboard(lang))

@dp.message(F.text.regexp(r"(Стандарт|Standart|Standard|VIP|Эксклюзив|Eksklüziv|Exclusive)"))
async def ticket_type_handler(message: types.Message):
    lang = user_lang.get(message.from_user.id, "en")
    
    ticket_type = None
    if message.text.startswith(("Стандарт", "Standart", "Standard")):
        ticket_type = "standard"
    elif message.text.startswith("VIP"):
        ticket_type = "vip"
    elif message.text.startswith(("Эксклюзив", "Eksklüziv", "Exclusive")):
        ticket_type = "exclusive"
    
    if not ticket_type:
        await message.answer("Неверный тип билета" if lang == "ru" else "Yanlış bilet növü" if lang == "az" else "Invalid ticket type")
        return
    
    user_data[message.from_user.id] = {
        "step": "name",
        "lang": lang,
        "ticket_type": ticket_type,
        "ticket_price": TICKET_TYPES[ticket_type][lang]["price"]
    }
    
    prompt = {
        "ru": "Для покупки билетов введите ваше Имя и Фамилию:",
        "az": "Bilet almaq üçün ad və soyadınızı daxil edin:",
        "en": "To buy tickets, please enter your First and Last name:"
    }[lang]
    
    await message.answer(prompt, reply_markup=types.ReplyKeyboardRemove())

@dp.message(lambda m: user_data.get(m.from_user.id, {}).get("step") == "name")
async def get_name(message: types.Message):
    user_data[message.from_user.id]["name"] = message.text
    user_data[message.from_user.id]["step"] = "phone"
    lang = user_data[message.from_user.id].get("lang", "en")
    
    prompt = {
        "ru": "Теперь введите ваш номер телефона:",
        "az": "İndi telefon nömrənizi daxil edin:",
        "en": "Now please enter your phone number:"
    }[lang]
    
    await message.answer(prompt)

@dp.message(lambda m: user_data.get(m.from_user.id, {}).get("step") == "phone")
async def get_phone(message: types.Message):
    phone = message.text
    if not phone.replace('+', '').isdigit() or len(phone) < 9:
        lang = user_data[message.from_user.id].get("lang", "en")
        error_msg = {
            "ru": "Пожалуйста, введите корректный номер телефона",
            "az": "Zəhmət olmasa, düzgün telefon nömrəsi daxil edin",
            "en": "Please enter a valid phone number"
        }[lang]
        await message.answer(error_msg)
        return
    
    user_data[message.from_user.id]["phone"] = phone
    user_data[message.from_user.id]["step"] = "confirm"
    lang = user_data[message.from_user.id].get("lang", "en")
    
    ticket_type = user_data[message.from_user.id]["ticket_type"]
    ticket_info = TICKET_TYPES[ticket_type][lang]
    
    confirmation = {
        "ru": f"Проверьте ваши данные:\n\n"
              f"🎟 Тип билета: {ticket_info['name']}\n"
              f"💳 Сумма: {ticket_info['price']}\n"
              f"👤 Имя: {user_data[message.from_user.id]['name']}\n"
              f"📱 Телефон: {phone}\n\n"
              f"Все верно?",
        "az": f"Məlumatlarınızı yoxlayın:\n\n"
              f"🎟 Bilet növü: {ticket_info['name']}\n"
              f"💳 Məbləğ: {ticket_info['price']}\n"
              f"👤 Ad: {user_data[message.from_user.id]['name']}\n"
              f"📱 Telefon: {phone}\n\n"
              f"Hər şey düzgündür?",
        "en": f"Please confirm your details:\n\n"
              f"🎟 Ticket type: {ticket_info['name']}\n"
              f"💳 Amount: {ticket_info['price']}\n"
              f"👤 Name: {user_data[message.from_user.id]['name']}\n"
              f"📱 Phone: {phone}\n\n"
              f"Is everything correct?"
    }[lang]
    
    keyboard = types.ReplyKeyboardMarkup(
        keyboard=[
            [types.KeyboardButton(text="✅ Да" if lang == "ru" else "✅ Bəli" if lang == "az" else "✅ Yes")],
            [types.KeyboardButton(text="❌ Нет" if lang == "ru" else "❌ Xeyr" if lang == "az" else "❌ No")]
        ],
        resize_keyboard=True
    )
    
    await message.answer(confirmation, reply_markup=keyboard)

@dp.message(F.text.in_(["✅ Да", "✅ Bəli", "✅ Yes"]))
async def confirm_purchase(message: types.Message):
    if message.from_user.id not in user_data:
        return
    
    lang = user_data[message.from_user.id].get("lang", "en")
    user_data[message.from_user.id]["step"] = "payment"
    
    payment_info = {
        "ru": f"Оплатите {user_data[message.from_user.id]['ticket_price']} на карту: `{PAYMENT_CARD}`\n"
              "и отправьте скриншот оплаты.",
        "az": f"{user_data[message.from_user.id]['ticket_price']} məbləğini kartla ödəyin: `{PAYMENT_CARD}`\n"
              "və ödəniş skrinşotu göndərin.",
        "en": f"Please pay {user_data[message.from_user.id]['ticket_price']} to card: `{PAYMENT_CARD}`\n"
              "and send payment screenshot."
    }[lang]
    
    await message.answer(payment_info, reply_markup=get_menu_keyboard(lang))

@dp.message(F.text.in_(["❌ Нет", "❌ Xeyr", "❌ No"]))
async def cancel_purchase(message: types.Message):
    lang = user_lang.get(message.from_user.id, "en")
    if message.from_user.id in user_data:
        del user_data[message.from_user.id]
    
    msg = {
        "ru": "Заказ отменен. Можете начать заново.",
        "az": "Sifariş ləğv edildi. Yenidən başlaya bilərsiniz.",
        "en": "Order canceled. You can start again."
    }[lang]
    
    await message.answer(msg, reply_markup=get_menu_keyboard(lang))

@dp.message(lambda m: user_data.get(m.from_user.id, {}).get("step") == "payment")
async def handle_payment(message: types.Message):
    lang = user_data[message.from_user.id].get("lang", "en")
    
    if message.photo:
        try:
            photo = message.photo[-1]
            file = await bot.get_file(photo.file_id)
            path = f"{PHOTOS_DIR}/{message.from_user.id}_{photo.file_id}.jpg"
            await bot.download_file(file.file_path, path)
            
            if save_to_excel(
                message.from_user.id,
                user_data[message.from_user.id]["name"],
                user_data[message.from_user.id]["phone"],
                user_data[message.from_user.id]["ticket_type"],
                user_data[message.from_user.id]["ticket_price"],
                path
            ):
                await notify_admin(
                    message.from_user.id,
                    user_data[message.from_user.id]["name"],
                    user_data[message.from_user.id]["phone"],
                    user_data[message.from_user.id]["ticket_type"],
                    user_data[message.from_user.id]["ticket_price"]
                )
                
                confirmation = {
                    "ru": "Спасибо! Ваша заявка на рассмотрении.",
                    "az": "Təşəkkürlər! Müraciətiniz nəzərdən keçirilir.",
                    "en": "Thank you! Your application is under review."
                }[lang]
                
                await message.answer(confirmation, reply_markup=get_menu_keyboard(lang))
                del user_data[message.from_user.id]
            
        except Exception as e:
            logger.error(f"Payment processing error: {e}")
            error_msg = {
                "ru": "Ошибка обработки платежа, попробуйте снова",
                "az": "Ödəniş emalı xətası, yenidən cəhd edin",
                "en": "Payment processing error, please try again"
            }[lang]
            await message.answer(error_msg)
    else:
        prompt = {
            "ru": "Пожалуйста, отправьте скриншот оплаты.",
            "az": "Zəhmət olmasa, ödəniş skrinşotu göndərin.",
            "en": "Please send the payment screenshot."
        }[lang]
        await message.answer(prompt)

# --- Admin Handlers ---
@dp.message(Command("admin"))
async def admin_command(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("⛔ Доступ запрещён!")
        return
        
    await message.answer(
        "🛠 *Панель администратора*",
        reply_markup=get_admin_keyboard(),
        parse_mode="Markdown"
    )

@dp.callback_query(F.data.startswith("admin_"))
async def handle_admin_callbacks(callback: types.CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Доступ запрещён!")
        return
    
    try:
        action = callback.data.split('_')[1]
        
        if action == "stats":
            report = await generate_stats_report()
            await callback.message.edit_text(report, reply_markup=get_admin_keyboard())
            
        elif action == "last_orders":
            orders = await get_last_orders()
            await callback.message.edit_text(orders, reply_markup=get_admin_keyboard())
            
        elif action == "search":
            await callback.message.answer("Введите ID пользователя:")
            admin_pending_actions[callback.from_user.id] = "waiting_for_id"
            
        elif action == "refresh":
            await callback.message.edit_text(
                "🛠 *Панель администратора*",
                reply_markup=get_admin_keyboard(),
                parse_mode="Markdown"
            )
            
        await callback.answer()
    except Exception as e:
        logger.error(f"Admin callback error: {e}")
        await callback.answer("⚠️ Произошла ошибка")

@dp.message(lambda m: admin_pending_actions.get(m.from_user.id) == "waiting_for_id")
async def handle_admin_search(message: types.Message):
    if not is_admin(message.from_user.id):
        return
        
    try:
        user_id = int(message.text)
        wb = openpyxl.load_workbook("tickets.xlsx")
        ws = wb.active
        
        found = None
        for row in ws.iter_rows(values_only=True):
            if row[0] == user_id:
                found = row
                break
                
        if not found:
            await message.answer("❌ Заявка не найдена")
        else:
            report = (
                f"🔍 *Найдена заявка:*\n\n"
                f"👤 *{found[1]}*\n"
                f"📞 `{found[2]}`\n"
                f"🎟 {found[3]} ({found[4]})\n"
                f"📸 [Фото]({found[5]})\n"
                f"🕒 {found[6]}"
            )
            await message.answer(report, parse_mode="Markdown")
            
    except ValueError:
        await message.answer("❌ Введите числовой ID")
    finally:
        admin_pending_actions.pop(message.from_user.id, None)

@dp.message(Command("accept"))
async def accept_request(message: types.Message):
    if not is_admin(message.from_user.id):
        return
        
    if not message.reply_to_message:
        await message.answer("ℹ️ Ответьте на сообщение с заявкой для подтверждения")
        return
        
    try:
        text = message.reply_to_message.text
        user_id = int(text.split("ID:")[1].split("\n")[0].strip())
        
        if user_id in pending_approvals:
            pending_approvals[user_id]["approved"] = True
            await message.answer(f"✅ Заявка {user_id} подтверждена")
            await bot.send_message(
                user_id,
                "🎉 Ваша заявка подтверждена! Билет активен."
            )
        else:
            await message.answer("⚠️ Заявка не найдена в ожидающих")
    except Exception as e:
        logger.error(f"Accept error: {e}")
        await message.answer("❌ Ошибка подтверждения")

@dp.message(Command("reject"))
async def reject_request(message: types.Message):
    if not is_admin(message.from_user.id):
        return
        
    if not message.reply_to_message:
        await message.answer("ℹ️ Ответьте на сообщение с заявкой для отклонения")
        return
        
    try:
        text = message.reply_to_message.text
        user_id = int(text.split("ID:")[1].split("\n")[0].strip())
        reason = message.text.split("/reject")[1].strip() if len(message.text.split("/reject")) > 1 else "не указана"
        
        if user_id in pending_approvals:
            pending_approvals[user_id]["approved"] = False
            
            await message.answer(f"❌ Заявка {user_id} отклонена")
            await bot.send_message(
                user_id,
                f"⚠️ Ваша заявка отклонена. Причина: {reason}"
            )
        else:
            await message.answer("⚠️ Заявка не найдена в ожидающих")
    except Exception as e:
        logger.error(f"Reject error: {e}")
        await message.answer("❌ Ошибка отклонения")

# --- Database Functions ---
def save_to_excel(user_id, name, phone, ticket_type, ticket_price, photo_path):
    try:
        file_path = "tickets.xlsx"
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["User ID", "Name", "Phone", "Ticket Type", "Ticket Price", "Photo Path", "Date"])
        
        ws.append([
            user_id,
            name,
            phone,
            ticket_type,
            ticket_price,
            photo_path,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ])
        wb.save(file_path)
        
        save_counter['total'] += 1
        if save_counter['total'] % 10 == 0:
            backup_path = f"tickets_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(backup_path)
            logger.info(f"Created backup: {backup_path}")
        
        return True
    except Exception as e:
        logger.error(f"Database error: {e}")
        return False

async def notify_admin(user_id: int, name: str, phone: str, ticket_type: str, ticket_price: str):
    try:
        msg = await bot.send_message(
            YOUR_TELEGRAM_ID,
            f"🆕 *Новая заявка на билет*\n\n"
            f"👤 ID: {user_id}\n"
            f"📛 Имя: {name}\n"
            f"📱 Телефон: `{phone}`\n"
            f"🎫 Тип: {ticket_type}\n"
            f"💵 Сумма: {ticket_price}\n"
            f"🕒 Время: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
            f"Ответьте на это сообщение командой:\n"
            f"/accept - подтвердить\n"
            f"/reject [причина] - отклонить",
            parse_mode="Markdown"
        )
        
        pending_approvals[user_id] = {
            "message_id": msg.message_id,
            "data": (user_id, name, phone, ticket_type, ticket_price),
            "approved": None
        }
    except Exception as e:
        logger.error(f"Failed to notify admin: {e}")

# --- Unhandled Messages ---
@dp.message()
async def handle_unmatched_messages(message: types.Message):
    if message.from_user.id == YOUR_TELEGRAM_ID:
        await message.answer("ℹ️ Используйте /admin для управления ботом")
    else:
        lang = user_lang.get(message.from_user.id, "en")
        response = {
            "ru": "Пожалуйста, используйте кнопки меню",
            "az": "Zəhmət olmasa menyu düymələrindən istifadə edin",
            "en": "Please use the menu buttons"
        }[lang]
        await message.answer(response, reply_markup=get_menu_keyboard(lang))

# --- Main ---
async def main():
    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)

if __name__ == "__main__":
    logging.info("Starting bot...")
    asyncio.run(main())